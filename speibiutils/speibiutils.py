import os
import sftp.sftp as sftpmodule
import pandas as pd
import logging
import openpyxl
import dotenv
from typing import List, Optional, Callable
import sys
import re
from datetime import date, datetime
import shutil
from config import MAX_BARCODES_LARGE, MAX_BARCODES_SMALL, MAX_DAYS_RETENTION, LARGE_TASK_HOUR, SBK_DIR

# Possible states of a task
STATES = ['NEW', 'READY', 'ERROR', 'PROCESSING', 'DONE']

# Possible sizes of a task
SIZE = ['SMALL', 'LARGE']


def sftp_connect(fn: Callable) -> Callable:
    """Decorator to connect to the SFTP server

    Parameters
    ----------
    fn : Callable
        Function to wrap

    Returns
    -------
    Callable
        Wrapped function
    """

    def wrapper(*args, **kwargs) -> Optional[any]:
        """Wrapper function to connect to the SFTP server

        Parameters
        ----------
        args : list
            List of arguments

        kwargs : dict
            Keyword arguments

        Returns
        -------
        Any
            Result of the function

        """
        dotenv.load_dotenv()
        host = os.getenv('SFTP_HOST')
        user = os.getenv('SFTP_USER')
        password = os.getenv('SFTP_PASSWORD')
        environment = os.getenv('SFTP_ENVIRONMENT')
        sftp = sftpmodule.SFTP(host, user, password)

        # We need to work in a particular directory for the test environment
        if environment == 'test':
            sftp.SFTP_Client.chdir(path=f'automation_storage_tasks')

        # Transmit sftp connection to the function
        kwargs['sftp'] = sftp

        # Start the function
        result = fn(*args, **kwargs)

        # Close the connection
        sftp.close()

        return result

    return wrapper


class Task:
    """Task class to handle tasks

    This class is mainly used to handle various paths to files and directories.

    Attributes
    ----------
    remote_path : str
        Remote path of the task
    """
    def __init__(self, remote_path: Optional[str] = None,
                 directory: Optional[str] = None,
                 account: Optional[str] = None) -> None:
        """Initialize a task object

        Remote path or directory + account must be provided

        Parameters
        ----------
        remote_path : str
            Remote path of the task

        directory : str
            Directory of the task

        account : str
            Account of the task

        Returns
        -------
        None
        """
        if remote_path is not None:
            self.remote_path = remote_path
        elif directory is not None and account is not None:
            self.remote_path = f'{account}/download/storage_tasks/{directory}'

    @staticmethod
    def get_name_from_dir(directory: str) -> Optional[str]:
        """Get the task name from the directory

        Remove STATE and SIZE from the directory name

        Parameters
        ----------
        directory : str
            Directory name containing the task

        Returns
        -------
        str
            Task name
        """
        if Task.is_valid_directory(directory) is False:
            return None

        return '_'.join(directory.split('_')[:-2])

    @staticmethod
    def is_valid_directory(directory) -> bool:
        """Check if a directory name is valid

        Parameters
        ----------
        directory : str
            Directory name

        Returns
        -------
        bool
            True if the directory is valid, False otherwise
        """
        m = re.match(r'^task_\d{4}-\d{2}-\d{2}_.*_([A-Z]+)_([A-Z]+)$',
                     directory)
        if m is None:
            return False

        if m.group(1) not in SIZE or m.group(2) not in STATES:
            return False

        return True

    def is_valid(self) -> bool:
        """Check if a task is valid

        Check entire path of the task

        Returns
        -------
        bool
            True if the task is valid, False otherwise
        """

        m = re.match(r'^(sbk\w+)/download/storage_tasks/task_\d{4}-\d{2}-\d{2}_.*_([A-Z]+)_([A-Z]+)$',
                     self.remote_path)
        if m is None:
            return False

        if m.group(1) not in SBK_DIR or m.group(2) not in SIZE or m.group(3) not in STATES:
            return False

        return True

    def get_name(self) -> Optional[str]:
        """Get the name of a task

        Returns
        -------
        str
            Name of the task without the state and size
        """
        if self.is_valid() is False:
            return None

        m = re.match(r'^sbk\w+/download/storage_tasks/(task_\d{4}-\d{2}-\d{2}_.*_[A-Z]+)_[A-Z]+$',
                     self.remote_path)

        return m.group(1)

    def get_directory(self) -> Optional[str]:
        """Get the directory of a task

        Returns
        -------
        str
            Directory of the task
        """
        if self.is_valid() is False:
            return None

        m = re.match(r'^sbk\w+/download/storage_tasks/(task_\d{4}-\d{2}-\d{2}_.*_[A-Z]+_[A-Z]+)$',
                     self.remote_path)

        return m.group(1)

    def get_directory_path(self, local: Optional[bool] = False) -> Optional[str]:
        """Get the directory path of a task

        Parameters
        ----------
        local : bool
            If True, return the local path, otherwise the remote path

        Returns
        -------
        str
            Directory path of the task (with state and size)
        """
        if self.is_valid() is False:
            return None

        m = re.match(r'^sbk\w+/download/storage_tasks/task_\d{4}-\d{2}-\d{2}_.*_[A-Z]+_[A-Z]+$',
                     self.remote_path)
        return f'data/{m.group(0)}' if local is True else m.group(0)

    def get_form_path(self, local: Optional[bool] = False) -> Optional[str]:
        """Get the form path of a task

        Parameters
        ----------
        local : bool
            If True, return the local path, otherwise the remote path

        Returns
        -------
        str
        """
        if self.is_valid() is False:
            return None

        directory_path = self.get_directory_path(local)
        task_name = self.get_name()
        return f'{directory_path}/{task_name}.xlsx'

    def get_form_name(self) -> Optional[str]:
        """Get the form file name of a Excel form

        Returns
        -------
        str
            Name of the form file
        """

        if self.is_valid() is False:
            return None

        return f'{self.get_name()}.xlsx'

    def get_processing_file_path(self, local: Optional[bool] = False) -> Optional[str]:
        """Get the processing file path of a task

        Parameters
        ----------
        local : bool
            If True, return the local path, otherwise the remote path

        Returns
        -------
        str
            Path of the processing file
        """
        if self.is_valid() is False:
            return None

        directory_path = self.get_directory_path(local)
        task_name = self.get_name()
        return f'{directory_path}/{task_name}_items_processing.csv'

    def get_scheduled_date(self) -> date:
        """Return the scheduled date in date format

        Returns
        -------
        date
            Scheduled date of the task
        """
        return datetime.strptime(self.get_parameters()['Scheduled_date'], '%Y-%m-%d').date()

    def get_parameters(self) -> Optional[dict]:
        """Get the parameters of a task

        Returns
        -------
        dict
            Parameters of the task, keys are: Account, Directory, Scheduled_date, Size, State
        """
        if self.is_valid() is False:
            return None

        m = re.match(r'^(sbk\w+)/download/storage_tasks/task_(\d{4}-\d{2}-\d{2})_.*_([A-Z]+)_([A-Z]+)$',
                     self.remote_path)

        if m is None:
            return

        return {
            'Account': m.group(1),
            'Directory': self.get_directory(),
            'Scheduled_date': m.group(2),
            'Size': m.group(3),
            'State': m.group(4)
        }

    def check_form_file(self, barcodes_from_other_tasks: Optional[list[str]] = None) -> (bool, List[str], List[str]):
        """Check if an Excel file is conform to the expected format

        Parameters
        ----------
        barcodes_from_other_tasks : list[str]
            List of barcodes from other tasks

        Returns
        -------
        bool
            True if the file is conform, False otherwise and a list of messages
        List[str]
            Barcodes list updated to avoid duplicates
        List[str]
            List of messages
        """
        if barcodes_from_other_tasks is None:
            barcodes_from_other_tasks = []
        messages = ['*** Checking Excel form conformity ***']

        try:
            wb = openpyxl.load_workbook(self.get_form_path(local=True))
        except Exception as e:
            error_message = f'Error reading file {self.get_form_name()}: {e}'
            logging.error(error_message)
            messages.append(error_message)
            return False, [], messages

        if wb.sheetnames != ['General', 'Items', 'Locations_mapping', 'Item_policies_mapping', 'data_validation']:
            error_message = (f"Bad or missing sheet names, must be ['General', 'Items', 'Locations_mapping', "
                             f"'Item_policies_mapping', 'data_validation']")
            logging.error(error_message)
            messages.append(error_message)
            return False, [], messages

        ws = wb['data_validation']

        version = ws.cell(row=2, column=4).value
        if version != os.getenv('SFTP_EXCEL_FORM_VERSION'):
            error_message = f'Version {version} not supported. Must be {os.getenv("SFTP_EXCEL_FORM_VERSION")}.'
            logging.error(error_message)
            messages.append(error_message)
            return False, [], messages
        messages.append(f'Excel file found: {self.get_form_name().split("/")[-1]}')
        messages.append(f'Version of Excel form supported: {version}')

        ws = wb['General']

        # iz_s = ws.cell(row=3, column=2).value
        iz_d = ws.cell(row=4, column=2).value
        env = {'Production': 'P',
               'Sandbox': 'S'}.get(ws.cell(row=5, column=2).value, 'P')

        size = self.get_parameters()['Size']

        if iz_d != 'VKSS' and env == 'P':
            error_message = f'Destination IZ "{iz_d}" selected, only IZ VKSS is allowed with automation.'
            logging.error(error_message)
            messages.append(error_message)
            return False, [], messages

        # Load barcodes
        if os.path.exists(self.get_processing_file_path(local=True)):
            # Process file already exists
            df = pd.read_csv(self.get_processing_file_path(local=True), dtype=str)
            df = df.replace('False', False)
            df = df.replace('True', True)
            df = df.replace('NaN', None)
            barcodes = df.loc[~df['Copied']]['Barcode'].tolist()
        else:
            # Load from Excel file
            barcodes = (pd.read_excel(self.get_form_path(local=True), sheet_name='Items', dtype=str)['Barcode']
                        .dropna().str.strip().str.strip("'").tolist())
        if size == 'LARGE' and len(barcodes) > MAX_BARCODES_LARGE:
            error_message = f'Too many barcodes ({len(barcodes)}), maximum is {MAX_BARCODES_LARGE}.'
            logging.error(error_message)
            messages.append(error_message)
            return False, [], messages
        elif size == 'LARGE' and len(barcodes) > MAX_BARCODES_SMALL:
            error_message = f'Too many barcodes ({len(barcodes)}), maximum is {MAX_BARCODES_SMALL}.'
            logging.error(error_message)
            messages.append(error_message)
            return False, [], messages
        elif len(barcodes) == 0:
            error_message = f'0 barcorde in the file'
            logging.error(error_message)
            messages.append(error_message)
            return False, [], messages

        elif len(set(barcodes) & set(barcodes_from_other_tasks)) > 0:
            error_message = f'Barcodes {set(barcodes) & set(barcodes_from_other_tasks)} already in other tasks'
            logging.error(error_message)
            messages.append(error_message)
            return False, [], messages

        elif len(barcodes) != len(set(barcodes)):
            seen = set()
            duplicates = set()

            for barcode in barcodes:
                if barcode in seen:
                    duplicates.add(barcode)
                else:
                    seen.add(barcode)

            error_message = f'Duplicate barcodes in the file: {duplicates}'
            logging.error(error_message)
            messages.append(error_message)
            return False, [], messages

        messages.append(f'{len(barcodes)} barcodes loaded from file.')

        # Load locations
        locations_table = pd.read_excel(self.get_form_path(local=True), sheet_name='Locations_mapping', dtype=str)
        if (len(locations_table) < 1 and
                locations_table.columns != ['Source library code', 'Source location code',
                                            'Destination library code', 'Destination location code']):
            error_message = f'Error with location table'
            logging.error(error_message)
            messages.append(error_message)
            return False, [], messages

        # Load item policies
        item_policies_table = pd.read_excel(self.get_form_path(local=True),
                                            sheet_name='Item_policies_mapping',
                                            dtype=str)
        if (len(item_policies_table) < 1 and
                item_policies_table.columns != ['Source item  code', 'Destination item policy code']):
            error_message = f'Error with item policies table'
            logging.error(error_message)
            messages.append(error_message)
            return False, [], messages

        logging.info('Excel file seems to be conform')
        messages.append('Excel file seems to be conform')
        return True, barcodes, messages


class LogFile:
    """Log file class to handle logs

    Logs can be linked to a task if a task name is provided.

    Attributes
    ----------
    file_name : str
        Name of the log file
    task : Task
        Task object

    logger : logging.Logger
            Logger object
    """

    def __init__(self, file_name: Optional[str] = "", task: Optional[Task] = None):
        """Initialize a log file object"""

        self.file_name = file_name
        self.task = task
        self.logger = self.config_log()

    def config_log(self) -> logging.Logger:
        """Set the log configuration for the entire process

        Returns
        -------
        logging.Logger
            Logger object
        """

        # Create the log folder if it doesn't exist
        if self.task is not None:
            log_path = self.task.get_directory_path(local=True)
        else:
            log_path = f'./data'
        if os.path.isdir(log_path) is False:
            os.mkdir(log_path)

        message_format = "%(asctime)s - %(levelname)s - %(message)s"
        log_file = logging.FileHandler(f'{log_path}/log{"" if len(self.file_name) == 0 else "_"}{self.file_name}.txt')

        # log_file.setFormatter(logging.Formatter(message_format))
        #
        # logger = logging.getLogger()  # root logger
        # logger.setLevel(logging.INFO)
        # for hdlr in logger.handlers[:]:  # remove all old handlers
        #     logger.removeHandler(hdlr)
        logging.basicConfig(format=message_format,
                            level=logging.INFO,
                            handlers=[log_file,
                                      logging.StreamHandler(sys.stdout)],
                            force=True)
        # logger.addHandler(log_file)
        # logger.addHandler(logging.StreamHandler(sys.stdout))
        return logging.getLogger()

    @staticmethod
    def close_log() -> None:
        """Close the log file"""
        logging.shutdown()


class TaskSummary:
    """Task summary class to handle the task list

    It uses in background a Excel file to store the task list. A copy is available locally and a copy in all folders
    of the SFTP accounts.

    Attributes
    ----------
    tasks : pd.DataFrame
        Task list
    """
    def __init__(self) -> None:
        """Initialize the task summary

        Load the task list from the Excel file

        Returns
        -------
        None
        """
        self.tasks = pd.read_excel('data/task_summary.xlsx', dtype=str).fillna('')

    @sftp_connect
    def update_task_state(self,
                          task: Task,
                          new_state: str,
                          sftp: sftpmodule.SFTP,
                          parameters: Optional[dict] = None) -> Optional[Task]:
        """Update the state of a task

        This method will update the state of a task and copy the task to the new state directory. It will also
        update the task summary, and the local + remote versions of the task.

        Parameters
        ----------
        task : Task
            Task object
        new_state : str
            New state of the task
        sftp : sftpmodule.SFTP
            SFTP connection
        parameters : dict
            Parameters to update, keys are: Account, Directory, Scheduled_date, Size, State

        Returns
        -------
        Task
            Updated task

        """
        logging.info(f'Updating task {task.get_name()} to state {new_state}')

        if parameters is None:
            parameters = {}

        if task.is_valid() is False:
            logging.error(f'Task {task.get_name()} update to new state impossible')
            return None

        new_task = Task(directory=self.update_task_name_state(task.get_directory(), new_state),
                        account=task.get_parameters()['Account'])

        # Copy to local version of the data with current state
        if os.path.exists(task.get_directory_path(local=True)) is False:
            sftp.copy_to_local(task.get_directory_path(), task.get_directory_path(local=True))

        # Once copied delete remote version
        sftp.rmtree(task.get_directory_path())

        # Remove local task if it already exists
        if os.path.exists(new_task.get_directory_path(local=True)) is True:
            logging.warning(f'{new_task.get_directory_path(local=True)} already exists => removing it')
            shutil.rmtree(new_task.get_directory_path(local=True))

        # Rename the local task to the new state
        os.rename(task.get_directory_path(local=True), new_task.get_directory_path(local=True))

        # Copy the local task to the remote server
        sftp.copy_to_remote(new_task.get_directory_path(local=True), new_task.get_directory_path())
        logging.info(f'{new_task.get_directory()} copied to remote server')

        # Update the task summary with a new task
        if len(self.tasks[self.tasks['Directory'] == task.get_directory()]) == 0:
            task_parameters = new_task.get_parameters()
            for p in parameters:
                task_parameters[p] = parameters[p]

            task_parameters['State'] = new_state
            self.tasks.loc[len(self.tasks)] = task_parameters
            logging.info(f'{task.get_directory()} added to the task summary')

        # Update the task summary with an existing task
        else:
            self.tasks.loc[self.tasks['Directory'] == task.get_directory(), 'State'] = new_state
            for p in parameters:
                self.tasks.loc[self.tasks['Directory'] == task.get_directory(), p] = parameters[p]

            self.tasks.loc[self.tasks['Directory'] == task.get_directory(), 'Directory'] = new_task.get_directory()
            logging.info(f'{task.get_directory()} updated in the task summary')

        logging.info(f'Task {task.get_directory()} updated to state {new_state}')
        self.save()
        return new_task

    @staticmethod
    def update_task_name_state(task_directory: str, new_state: str) -> Optional[str]:
        """Update the state of a task name

        Parameters
        ----------
        task_directory : str
            Task directory name
        new_state : str
            New state of the task

        Returns
        -------
        str
            New task name with the new state
        """
        if not Task.is_valid_directory(task_directory):
            return

        return re.sub(r'_[A-Z]+$', f'_{new_state}', task_directory)

    def get_directories(self) -> List[str]:
        """Get the directories from the task list

        Returns
        -------
        List[str]
            List of directories"""

        return self.tasks['Directory'].tolist()

    def is_task_date_available(self, task: Task) -> (bool, str):
        """Check if a task date is available

        Parameters
        ----------
        task : Task
            Task object

        Returns
        -------
        bool
            True if the task date is available, False otherwise
        str
            Error message if the task date is not available
        """
        task_parameters = task.get_parameters()
        existing_tasks = self.tasks.loc[((self.tasks['Scheduled_date'] == task_parameters['Scheduled_date'])
                                         & (self.tasks["Size"] == task_parameters['Size']) &
                                         (self.tasks['State'].isin(['NEW', 'READY', 'PROCESSING'])))]

        if task_parameters['Size'] == 'LARGE':
            if len(existing_tasks) >= 1:
                return False, 'A large task is already scheduled for this date'
            if date.today().isoformat() == task_parameters['Scheduled_date'] and datetime.now().hour > LARGE_TASK_HOUR:
                return False, 'Large tasks must be scheduled before 18:00 if the date is today'

        elif task_parameters['Size'] == 'SMALL':
            if len(existing_tasks) > 3:
                return False, 'Too many small tasks are already scheduled for this date'
            if (date.today().isoformat() == task_parameters['Scheduled_date']
                    and datetime.now().hour + len(existing_tasks) > LARGE_TASK_HOUR-1):
                return False, 'Too many small tasks are already scheduled for this date'

        return True, None

    @sftp_connect
    def save(self, sftp: sftpmodule.SFTP) -> None:
        """Save the task summary

        The summary is saved locally and distributed on all accounts remotely.

        Parameters
        ----------
        sftp : sftpmodule.SFTP
            SFTP connection

        Returns
        -------
        None"""
        self.tasks.to_excel('./data/task_summary.xlsx', index=False)
        for directory in SBK_DIR:
            sftp.put('./data/task_summary.xlsx',
                     f'./{directory}/download/storage_tasks/task_summary.xlsx')

    @staticmethod
    def clean_local_directories() -> None:
        """Clean the local directories

        Remove outdated directories

        Returns
        -------
        None
        """

        for account in SBK_DIR:
            if os.path.isdir(f'./data/{account}') is False:
                os.mkdir(f'./data/{account}')
            if os.path.isdir(f'./data/{account}/download') is False:
                os.mkdir(f'./data/{account}/download')
            if os.path.isdir(f'./data/{account}/download/storage_tasks') is False:
                os.mkdir(f'./data/{account}/download/storage_tasks')
            if os.path.isdir(f'./data/{account}/upload/') is False:
                os.mkdir(f'./data/{account}/upload/')
            if os.path.isdir(f'./data/{account}/upload/storage_tasks') is False:
                os.mkdir(f'./data/{account}/upload/storage_tasks')

            for directory in os.listdir(f'./data/{account}/download/storage_tasks'):
                temp_task = Task(directory=directory, account=account)
                if (date.today() - temp_task.get_scheduled_date()).days > MAX_DAYS_RETENTION:
                    shutil.rmtree(temp_task.get_directory_path(local=True))
                    logging.warning(f'Local directory {temp_task.get_directory_path(local=True)}'
                                    f'{directory} too old => removing it')

    @sftp_connect
    def clean_remote_directories(self, sftp: sftpmodule.SFTP) -> None:
        """Clean the directories

        Parameters
        ----------
        sftp : sftpmodule.SFTP
            SFTP connection

        Returns
        -------
        None
        """
        remote = RemoteLocation()
        self.tasks = self.tasks.loc[self.tasks['Directory'].isin(remote.directories)]

        # 1st run to clean outdated tasks, bad naming, duplicate
        for entry_path in remote.paths:

            task = Task(entry_path)

            # Delete tasks with bad name
            if task.is_valid() is False:
                logging.error(f'{task.get_directory()} is not a valid task name => removing it')
                sftp.rmtree(entry_path)
                continue

            # Remove old remote entries
            if (date.today() - task.get_scheduled_date()).days > MAX_DAYS_RETENTION:
                logging.warning(f'{task.get_directory()} is too old => removing it')
                sftp.rmtree(entry_path)
                continue

            # No same task name is allowed, task in ERROR state can be overwritten and are ignored
            if (task.get_name() in
                    self.tasks.loc[self.tasks['State'] != 'ERROR']['Directory'].apply(task.get_name_from_dir).values):
                logging.error(f'{task.get_directory()}: workflow broken, same task name already exists')
                sftp.rmtree(entry_path)
                continue

        self.save()

        # 2nd run to update the task summary
        remote = RemoteLocation()

        self.tasks = self.tasks.loc[self.tasks['Directory'].isin(remote.directories)]

        for entry_path in remote.paths:

            task = Task(entry_path)

            # No need to check the task if it is already in the task summary
            if task.get_directory() in self.get_directories():
                continue

            task_parameters = task.get_parameters()

            # Only new tasks can be added to the task summary
            if task_parameters['State'] not in ['NEW', 'ERROR']:
                error_message = (f'{task.get_directory()}: workflow broken, new tasks must have '
                                 f'"NEW" state and not "{task_parameters["State"]}"')
                logging.error(error_message)
                self.update_task_state(task, new_state='ERROR', parameters={'Message': error_message})
                continue

            # Check if the task date is available
            date_validity, error_message = self.is_task_date_available(task)
            if date_validity is False:
                logging.error(error_message)
                self.update_task_state(task, new_state='ERROR', parameters={'Message': error_message})
                continue

            self.tasks.loc[task.get_directory()] = task_parameters
            # logging.info(f'New task found: {remote_directory}')

        self.save()

    @sftp_connect
    def check_forms_conformity(self, sftp: sftpmodule.SFTP) -> None:
        """Check if the forms are conform

        Parameters
        ----------
        sftp : sftpmodule.SFTP
            SFTP connection

        Returns
        -------
        None
        """

        entries = self.tasks.loc[self.tasks['State'] == 'NEW', ['Account', 'Directory']].values

        if len(entries) == 0:
            logging.warning('No task found with state "NEW" existing in the task summary')
            return

        barcodes_from_other_tasks = []
        for account, directory in entries:
            task = Task(directory=directory, account=account)

            if sftp.is_file(task.get_form_path()) is False:
                error_message = f'Missing file {task.get_form_name()}'
                logging.error(error_message)
                self.update_task_state(task, new_state='ERROR',
                                       parameters={'Message': error_message})
                continue

            sftp.copy_to_local(task.get_directory_path(), task.get_directory_path(local=True))
            logging.info(f'{directory} copied to local server')

            is_conform, barcodes, messages = task.check_form_file(barcodes_from_other_tasks)

            barcodes_from_other_tasks += barcodes
            with open(f'{task.get_directory_path(local=True)}/form_check_result.txt', 'w') as f:
                f.write('\n'.join(messages))

            if is_conform is False:
                self.update_task_state(task,
                                       new_state='ERROR',
                                       parameters={'Message': f'{directory}: invalid form',
                                                   'Check_time': f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'})

            else:
                self.update_task_state(task,
                                       new_state='READY',
                                       parameters={'Check_time': datetime.now().strftime("%Y-%m-%d %H:%M:%S")})

    def get_processing_task(self) -> Optional[Task]:
        """Check if a processing task exists

        Returns
        -------
        str
            Remote path to the processing task, None if no current processing task
        """
        processing_tasks = self.tasks.loc[self.tasks['State'] == 'PROCESSING']
        if len(processing_tasks) == 0:
            return None
        processing_task = processing_tasks.iloc[0]

        return Task(account=processing_task["Account"], directory=processing_task["Directory"])

    @sftp_connect
    def get_next_task(self, size: str, sftp: sftpmodule.SFTP) -> Optional[Task]:
        """Get the next task to process

        Parameters
        ----------
        size : str
            Size of the task

        sftp : sftpmodule.SFTP
            SFTP connection

        Returns
        -------
        Optional[str]
            Remote path of the next task, None if no task is available
        """
        next_tasks = self.tasks.loc[(self.tasks['State'] == 'READY') &
                                    (self.tasks['Size'] == size) &
                                    (self.tasks['Scheduled_date'] == date.today().isoformat())]

        if len(next_tasks) == 0:
            logging.warning(f'No {size} task to process at the moment')
            return None

        next_task = Task(directory=next_tasks.iloc[0, 1], account=next_tasks.iloc[0, 0])

        sftp.copy_to_local(next_task.get_directory_path(), next_task.get_directory_path(local=True))
        logging.info(f'{next_task.get_directory()} copied to local server')

        if os.path.isfile(next_task.get_form_path(local=True)) is False:
            error_message = (f'Missing file {next_task.get_form_name()}'
                             f' => not possible to process the task')
            logging.error(error_message)
            self.update_task_state(next_task, new_state='ERROR',
                                   parameters={'Message': error_message})
            return None

        is_conform, _, messages = next_task.check_form_file()

        with open(f'data/{next_task.get_directory_path()}/form_check_result.txt', 'w') as f:
            f.write('\n'.join(messages))

        if is_conform is False:
            error_message = (f'Form not conform: {next_task.get_name()} => impossible '
                             f' => not possible to process the task')
            logging.error(error_message)
            self.update_task_state(next_task,
                                   new_state='ERROR',
                                   parameters={'Message': error_message,
                                               'Check_time': f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'})
            return None

        return next_task


class NewTask:
    def __init__(self, form_path: str) -> None:
        self.form_path = form_path
        self.error = self.is_valid_form_path(form_path) is False

        if self.form_path.endswith('DELETE.xlsx') is True:
            self.delete_task()
        elif self.form_path.endswith('RESTART.xlsx') is True:
            self.restart_task()
        elif self.error is False:
            self.create_new_task()

    @staticmethod
    def is_valid_form_path(form_path: str) -> bool:
        """Check if a task name is valid

        Parameters
        ----------
        form_path : str
            Task name

        Returns
        -------
        bool
            True if the task name is valid, False otherwise
        """
        m = re.match(r'^sbk\w+/upload/storage_tasks/task_\d{4}-\d{2}-\d{2}_.*_(?:SMALL|LARGE)(?:_DELETE|_RESTART)?\.xlsx$',
                     form_path)
        if m is None:
            return False

        return True

    def get_directory(self) -> str:
        """Get the directory of a task

        Returns
        -------
        str
            Directory of the task
        """
        m = re.match(r'^(sbk\w+)/upload/storage_tasks/task_\d{4}-\d{2}-\d{2}_.*_(?:SMALL|LARGE)(?:_DELETE|_RESTART)?\.xlsx$',
                     self.form_path)

        return m.group(1)

    def get_form_name(self) -> Optional[str]:
        """Get the form file name of a Excel form

        Returns
        -------
        str
            Name of the form file
        """
        m = re.match(r'^sbk\w+/upload/storage_tasks/(task_\d{4}-\d{2}-\d{2}_.*_(?:SMALL|LARGE)(?:_DELETE|_RESTART)?\.xlsx)$',
                     self.form_path)

        return m.group(1)

    def get_task_name(self, state: Optional[str] = None) -> str:
        """Get the name of a task

        Parameters
        ----------
        state : str
            State of the task, useful to create path of the tasks
        """

        m = re.match(r'^sbk\w+/upload/storage_tasks/(task_\d{4}-\d{2}-\d{2}_.*_(?:SMALL|LARGE))(?:_DELETE|_RESTART)?\.xlsx$',
                     self.form_path)
        if state is None:
            return m.group(1)
        else:
            return f'{m.group(1)}_{state}'

    @sftp_connect
    def delete_task(self, sftp: sftpmodule.SFTP) -> None:
        """Delete a task
        """
        for directory in sftp.listdir(f'{self.get_directory()}/download/storage_tasks/'):
            if directory in [self.get_task_name(state='NEW'), self.get_task_name(state='READY')]:
                sftp.rmtree(f'{self.get_directory()}/download/storage_tasks/{directory}')
                sftp.remove(self.form_path)
                logging.info(f'{self.get_task_name()} deleted')
                return

    @sftp_connect
    def restart_task(self, sftp: sftpmodule.SFTP) -> None:
        """Restart a task
        """
        m = re.match(r'^sbk\w+/upload/storage_tasks/task_(\d{4}-\d{2}-\d{2})_(\d{4}-\d{2}-\d{2})(.*)_(SMALL|LARGE)_RESTART\.xlsx$',
                     self.form_path)
        if m is None:
            logging.error(f'{self.form_path}: Invalid form name')
            return
        current_task_dir = f'{self.get_directory()}/download/storage_tasks/task_{m.group(1)}{m.group(3)}_{m.group(4)}_DONE'
        new_task_dir = f'{self.get_directory()}/download/storage_tasks/task_{m.group(2)}{m.group(3)}_SMALL_NEW'

        if sftp.is_dir(current_task_dir) is True:
            sftp.rename(current_task_dir, new_task_dir)
            for f in sftp.listdir(new_task_dir):
                if f.endswith('_LARGE.xlsx'):
                    sftp.rename(f'{new_task_dir}/{f}',
                                f'{new_task_dir}/task_{m.group(2)}{m.group(3)}_SMALL.xlsx')
                elif f.endswith('_LARGE_items_processing.csv'):
                    sftp.rename(f'{new_task_dir}/{f}',
                                f'{new_task_dir}/task_{m.group(2)}{m.group(3)}_SMALL_items_processing.csv')
                elif f.endswith('LARGE_items_not_copied.csv'):
                    sftp.rename(f'{new_task_dir}/{f}',
                                f'{new_task_dir}/task_{m.group(2)}{m.group(3)}_SMALL_items_not_copied.csv')
            sftp.remove(self.form_path)
            logging.info(f'{self.get_task_name(state="NEW")} restarted')

        else:
            logging.error(f'{self.form_path}: Task not found')

    @sftp_connect
    def create_new_task(self, sftp: sftpmodule.SFTP) -> None:
        """Create a new task
        """
        for directory in sftp.listdir(f'{self.get_directory()}/download/storage_tasks/'):
            if self.get_task_name() in directory:

                # Suppress matching tasks with error state
                if directory.endswith('_ERROR'):
                    sftp.rmtree(f'{self.get_directory()}/download/storage_tasks/{directory}')
                else:
                    logging.error(f'{self.get_task_name()} already exists')
                    return

        sftp.mkdir(f'{self.get_directory()}/download/storage_tasks/{self.get_task_name(state="NEW")}')

        sftp.rename(self.form_path,
                    f'{self.get_directory()}/download/storage_tasks/'
                    f'{self.get_task_name(state="NEW")}/{self.get_form_name()}')


class RemoteLocation:
    """Remote location class to list the remote directories

    Attributes
    ----------
    paths : List[str]
        List of remote paths
    directories : List[str]
        List of directories"""

    def __init__(self) -> None:
        """Initialize the remote location

        Returns
        -------
        None
        """
        self.paths = self.get_remote_directories()
        self.directories = [p.split('/')[-1] for p in self.paths]

    @sftp_connect
    def get_remote_directories(self, sftp: sftpmodule.SFTP) -> List[str]:
        """Get the remote directories

        Parameters
        ----------
        sftp : sftpmodule.SFTP
            SFTP connection

        Returns
        -------
        List[str]
            List of directories
        """
        remote_directories = []
        for account_directory in SBK_DIR:
            if sftp.is_dir(f'./{account_directory}/download/storage_tasks') is False:
                logging.warning(f'Creating directory {account_directory}/download/storage_tasks')
                sftp.mkdir(f'./{account_directory}/download/storage_tasks')

            for entry in sftp.listdir(f'./{account_directory}/download/storage_tasks'):
                if sftp.is_dir(f'./{account_directory}/download/storage_tasks/{entry}') is True:
                    remote_directories += [f'{account_directory}/download/storage_tasks/{entry}']

        return remote_directories

    @sftp_connect
    def get_new_tasks(self, sftp: sftpmodule.SFTP) -> List[str]:
        """Get the new forms

        Returns
        -------
        List[str]
            List of new forms
        """
        new_tasks = []
        for account_directory in SBK_DIR:
            if sftp.is_dir(f'./{account_directory}/upload/storage_tasks') is False:
                sftp.mkdir(f'./{account_directory}/upload/storage_tasks')
            for entry in sftp.listdir(f'./{account_directory}/upload/storage_tasks'):
                if NewTask.is_valid_form_path(f'{account_directory}/upload/storage_tasks/{entry}'):
                    new_tasks.append(f'{account_directory}/upload/storage_tasks/{entry}')

        return new_tasks


if __name__ == '__main__':
    pass
